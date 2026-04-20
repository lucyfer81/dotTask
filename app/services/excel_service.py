from datetime import datetime, date as date_type
from openpyxl import Workbook
from openpyxl import load_workbook
from app import db
from app.models import Location, Task, TaskAssignment


LOCATION_HEADERS = [
    "Location ID", "Location Name", "Country", "City", "Location Type",
    "Region", "Is Active", "IT Manager", "Primary IT Contact",
]

TASK_HEADERS = [
    "Task ID", "Task Name", "Task Source", "Stakeholder", "Task Description",
    "Scope Type", "Scope Rule", "Scope Detail", "Task Owner", "Execution Model",
    "Overall Status", "Start Date", "Target Date", "Last Update",
    "Link to File / Collection", "Link to Mail", "Task Priority", "Comments",
]

ASSIGNMENT_HEADERS = [
    "Task ID", "Location ID", "IT Name", "IT Role", "Local Responsibility",
    "Local Status", "Last Update", "Issue / Blocker", "Comments",
]


def export_to_workbook():
    wb = Workbook()

    # Task_Master
    ws_task = wb.active
    ws_task.title = "Task_Master"
    ws_task.append(TASK_HEADERS)
    for t in Task.query.order_by(Task.id):
        ws_task.append([
            t.id, t.task_name, t.task_source, t.stakeholder, t.task_description,
            t.scope_type, t.scope_rule, t.scope_detail, t.task_owner, t.execution_model,
            t.overall_status, str(t.start_date) if t.start_date else "",
            str(t.target_date) if t.target_date else "",
            str(t.last_update) if t.last_update else "",
            t.link_to_file, t.link_to_mail, t.task_priority, t.comments,
        ])

    # Location_Master
    ws_loc = wb.create_sheet("Location_Master")
    ws_loc.append(LOCATION_HEADERS)
    for loc in Location.query.order_by(Location.id):
        ws_loc.append([
            loc.id, loc.location_name, loc.country, loc.city, loc.location_type,
            loc.region, "Yes" if loc.is_active else "No",
            loc.it_manager, loc.primary_it_contact,
        ])

    # Task_Assignment
    ws_assign = wb.create_sheet("Task_Assignment")
    ws_assign.append(ASSIGNMENT_HEADERS)
    for a in TaskAssignment.query.order_by(TaskAssignment.id):
        ws_assign.append([
            a.task_id, a.location_id, a.it_name, a.it_role, a.local_responsibility,
            a.local_status, str(a.last_update) if a.last_update else "",
            a.issue_blocker, a.comments,
        ])

    return wb


def import_from_workbook(wb, sheet_name=None):
    """Import Excel data. Returns stats {sheet_name: count}."""
    stats = {}

    if sheet_name is None or sheet_name == "Location_Master":
        if "Location_Master" in wb.sheetnames:
            ws = wb["Location_Master"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]:
                    continue
                loc = Location(
                    location_name=str(row[1]),
                    country=str(row[2] or ""),
                    city=str(row[3] or ""),
                    location_type=str(row[4] or ""),
                    region=str(row[5] or ""),
                    is_active=str(row[6]).lower() not in ("no", "false", "0"),
                    it_manager=str(row[7] or ""),
                    primary_it_contact=str(row[8] or ""),
                )
                db.session.add(loc)
                count += 1
            stats["Location_Master"] = count

    if sheet_name is None or sheet_name == "Task_Master":
        if "Task_Master" in wb.sheetnames:
            ws = wb["Task_Master"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]:
                    continue
                task = Task(
                    task_name=str(row[1]),
                    task_source=str(row[2] or ""),
                    stakeholder=str(row[3] or ""),
                    task_description=str(row[4] or ""),
                    scope_type=str(row[5] or "Manual"),
                    scope_rule=str(row[6] or ""),
                    scope_detail=str(row[7] or ""),
                    task_owner=str(row[8] or ""),
                    execution_model=str(row[9] or ""),
                    overall_status=str(row[10] or "Not Started"),
                    start_date=_to_date(row[11]),
                    target_date=_to_date(row[12]),
                    last_update=_to_date(row[13]),
                    link_to_file=str(row[14] or ""),
                    link_to_mail=str(row[15] or ""),
                    task_priority=str(row[16] or "Medium"),
                    comments=str(row[17] or ""),
                )
                db.session.add(task)
                count += 1
            stats["Task_Master"] = count

    db.session.commit()
    return stats


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_type):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None
